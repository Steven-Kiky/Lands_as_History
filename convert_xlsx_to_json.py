import openpyxl
import json
import re
import sys

# ========== Erected By -> Category mapping ==========
ERECTED_BY_CATEGORY = {
    # Ann Arbor Historical Foundation
    'Ann Arbor Historical Foundation': 'Ann Arbor Historical Foundation',
    'Ann Arbor Historical Foundation, University of Michigan': 'Ann Arbor Historical Foundation',

    # University of Michigan
    'University of Michigan': 'University of Michigan',
    'University of Michigan ': 'University of Michigan',
    'History and Traditions Committee': 'University of Michigan',
    'History and Tradition Committee': 'University of Michigan',
    'History and Traditions Committee, University of Michigan': 'University of Michigan',
    'History and Traditions Committee, Paid Advertisement; University of Michigan': 'University of Michigan',
    'University of Michigan History and Traditions Committee': 'University of Michigan',
    'The Regents of the University of Michigan': 'University of Michigan',
    'Matthaei Botanical Gardens and Nichols Arboretum, University of Michigan': 'University of Michigan',
    'University of Michigan Matthaei Botanical Gardens and Nichols Arboretum': 'University of Michigan',
    'University of Michigan Matthaei Botanical Gardens and Nichols Arboretum.': 'University of Michigan',
    'Ross School of Business and Planet Blue (University of Michigan\'s Sustainability Initiative)': 'University of Michigan',
    'Class of 1899, University of Michigan': 'University of Michigan',

    # State of Michigan
    'Michigan History Division, Department of State': 'State of Michigan',
    'Michigan History Division, Department of State. (Marker Number S494.)': 'State of Michigan',
    'Bureau of History, Michigan Department of State': 'State of Michigan',
    'Bureau of History, Michigan Department of State. (Marker Number L1181.)': 'State of Michigan',
    'Michigan Historical Commission': 'State of Michigan',
    'Michigan Historical Commission, Michigan History Center': 'State of Michigan',
    'Michigan Historical Commission - Michigan History Center': 'State of Michigan',
    'Michigan Historical Commission and the Michigan Historical Center': 'State of Michigan',
    'Michigan Historical Center and Michigan Department of State': 'State of Michigan',
    'Michigan Historical Center, Michigan Department of State': 'State of Michigan',
    'Historical Society of Michigan': 'State of Michigan',

    # Washtenaw County
    'Washtenaw County Historic District Commission': 'Washtenaw County',
    'Washtenaw County Historic District Commission.': 'Washtenaw County',
    'Washtenaw County Historic Distric Commission': 'Washtenaw County',
    'Washtenaw County Historical District Commission': 'Washtenaw County',
    'Washtenaw County Parks and Recreation Commission': 'Washtenaw County',
    'Parker Mill County Park Washtenaw County Parks and Recreation Commission': 'Washtenaw County',
    'Parker Mill County Park, Washtenaw County Parks and Recreation Commission': 'Washtenaw County',
    'Parker Mill County Park, Washtenaw County Park and Recreation Commission.': 'Washtenaw County',

    # Ann Arbor Historic District Commission
    'Ann Arbor Historic District Commission': 'Ann Arbor Historic District Commission',
    'Ann Arbor Historic District Commission and South University Area Association': 'Ann Arbor Historic District Commission',

    # Pittsfield Township
    'Pittsfield Township Historical Society': 'Pittsfield Township',
    'Pittsfield Charter Township': 'Pittsfield Township',
    'Pittsfield Charter Township Historic District Commission': 'Pittsfield Township',

    # Fraternal Organizations
    'National Council of Acacia': 'Fraternal Organizations',
    'Brothers of the Peninsular Chapter of Alpha Delta Phi': 'Fraternal Organizations',
    'Delta Sigma Delta': 'Fraternal Organizations',
    'Xi Psi Phi': 'Fraternal Organizations',

    # Private Donors
    'Family and friends of Charles Dension ': 'Private Donors',
    'Family and friends of Charles Dension': 'Private Donors',
    'Hazel and Edmund Koli': 'Private Donors',
    'Carol and Robert Mull': 'Private Donors',
    'Elizabeth R. Dean Fund and other contributions': 'Private Donors',

    # Patriotic Societies
    'Sons of the American Revolution': 'Patriotic Societies',
    'Daughters of the American Revolution': 'Patriotic Societies',
    'Sarah Caswell Angell Chapter and the Ypsilanti Chapter Daughters of the American Revolution': 'Patriotic Societies',
    'NSSAR George Washington Endowment Fund, MISSAR, NSDAR': 'Patriotic Societies',

    # DTE Energy
    'DTE Energy': 'DTE Energy',

    # City of Ann Arbor
    'City of Ann Arbor': 'City of Ann Arbor',
    'City of Ann Arbor Parks and Recreation': 'City of Ann Arbor',

    # Professional Societies
    'American Physical Society': 'Professional Societies',

    # Federal Government
    'United States Department of the Interior': 'Federal Government',

    # Community Organizations
    'Historical Marker Database': 'Community Organizations',

    # Unknown
    'N/A': 'Unknown',
}


def parse_location(loc_str):
    """Parse location string like '42° 16.774' N, 83° 44.635' W' into (lat, lng)."""
    if not loc_str or loc_str.strip() == 'N/A':
        return None, None

    loc_str = loc_str.strip()

    # Try DMS format: 42°16'41.1"N 83°44'17.8"W
    dms_pattern = r"""(\d+)[°]\s*(\d+)[''′]\s*([\d.]+)[""″]?\s*([NS])\s*[,\s]+\s*(\d+)[°]\s*(\d+)[''′]\s*([\d.]+)[""″]?\s*([EW])"""
    m = re.search(dms_pattern, loc_str)
    if m:
        lat = int(m.group(1)) + int(m.group(2)) / 60 + float(m.group(3)) / 3600
        if m.group(4) == 'S':
            lat = -lat
        lng = int(m.group(5)) + int(m.group(6)) / 60 + float(m.group(7)) / 3600
        if m.group(8) == 'W':
            lng = -lng
        return round(lat, 6), round(lng, 6)

    # Try degree-decimal-minutes format: 42° 16.774' N, 83° 44.635' W
    # Also handles Unicode degree signs and various quote marks
    ddm_pattern = r"""(\d+)\s*[°\u00b0]\s*([\d.]+)\s*[''′\u2019]?\s*([NS])\s*[,\s]+\s*(\d+)\s*[°\u00b0]\s*([\d.]+)\s*[''′\u2019]?\s*([EW])"""
    m = re.search(ddm_pattern, loc_str)
    if m:
        lat = int(m.group(1)) + float(m.group(2)) / 60
        if m.group(3) == 'S':
            lat = -lat
        lng = int(m.group(4)) + float(m.group(5)) / 60
        if m.group(6) == 'W':
            lng = -lng
        return round(lat, 6), round(lng, 6)

    # Try decimal degrees: 42.279567, -83.743917
    dd_pattern = r"""([-\d.]+)\s*[,\s]+\s*([-\d.]+)"""
    m = re.search(dd_pattern, loc_str)
    if m:
        lat = float(m.group(1))
        lng = float(m.group(2))
        if abs(lat) <= 90 and abs(lng) <= 180:
            return round(lat, 6), round(lng, 6)

    return None, None


def parse_year(date_val):
    """Parse year from various formats."""
    if date_val is None or str(date_val).strip() == '' or str(date_val).strip() == 'N/A':
        return None
    if isinstance(date_val, (int, float)):
        return int(date_val)
    # Try to extract year from string
    m = re.search(r'(\d{4})', str(date_val))
    if m:
        return int(m.group(1))
    return None


def get_category(erected_by):
    """Map erected_by to a category."""
    if not erected_by or erected_by.strip() == '':
        return 'Unknown'
    eb = erected_by.strip()
    if eb in ERECTED_BY_CATEGORY:
        return ERECTED_BY_CATEGORY[eb]
    # Try fuzzy matching for entries with special characters
    for key, val in ERECTED_BY_CATEGORY.items():
        if key.lower().strip() == eb.lower().strip():
            return val
    # Check if it contains known keywords
    eb_lower = eb.lower()
    if 'michigan historical' in eb_lower or 'michigan history' in eb_lower or 'department of state' in eb_lower:
        return 'State of Michigan'
    if 'university of michigan' in eb_lower:
        return 'University of Michigan'
    if 'washtenaw county' in eb_lower:
        return 'Washtenaw County'
    if 'ann arbor historic' in eb_lower:
        return 'Ann Arbor Historic District Commission'
    if 'pittsfield' in eb_lower:
        return 'Pittsfield Township'
    print(f'  WARNING: Unknown erected_by category: {repr(eb)}', file=sys.stderr)
    return 'Unknown'


def main():
    wb = openpyxl.load_workbook(r'd:\aaa umich2\lands\Historical Markers.xlsx')
    ws = wb['Raw Data']

    markers = []
    failed_coords = []

    for r in range(2, 201):  # 199 data rows
        title = ws.cell(row=r, column=1).value
        if not title or not str(title).strip():
            break

        title = str(title).strip()
        primary = str(ws.cell(row=r, column=2).value or '').strip()
        secondary = str(ws.cell(row=r, column=3).value or '').strip()
        inscription = str(ws.cell(row=r, column=4).value or '').strip()
        location = str(ws.cell(row=r, column=5).value or '').strip()
        date_val = ws.cell(row=r, column=6).value
        erected_by = str(ws.cell(row=r, column=7).value or '').strip()

        # Clean up
        if secondary == 'N/A' or secondary == 'None':
            secondary = ''

        lat, lng = parse_location(location)
        year = parse_year(date_val)
        category = get_category(erected_by if erected_by != 'N/A' else 'N/A')

        marker_id = r - 2  # 0-based ID

        if lat is None or lng is None:
            failed_coords.append((marker_id, title, location))

        marker = {
            "id": marker_id,
            "title": title,
            "primary": primary,
            "secondary": secondary,
            "lat": lat,
            "lng": lng,
            "year": year,
            "erected_by": erected_by,
            "erected_by_category": category,
            "inscription": inscription
        }
        markers.append(marker)

    # Write JSON
    with open(r'd:\aaa umich2\lands\markers.json', 'w', encoding='utf-8') as f:
        json.dump(markers, f, ensure_ascii=False, indent=2)

    print(f'Successfully converted {len(markers)} markers to markers.json')

    if failed_coords:
        print(f'\nWARNING: {len(failed_coords)} markers with unparseable coordinates:')
        for mid, title, loc in failed_coords:
            print(f'  ID {mid}: {title} -> {repr(loc)}')


if __name__ == '__main__':
    main()
